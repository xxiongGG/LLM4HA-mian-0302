import sys
from tqdm import tqdm
from openpyxl import load_workbook
from pygoogletranslation import Translator


def Google_Translate(file_home, out_home):
    wb = load_workbook(filename=file_home)
    ws = wb['Sheet2']
    translator = Translator(service_url='https://translate.google.com.hk/?hl=zh-CN')

    nrows = ws.max_row
    ncols = ws.max_column
    print('File rows num is:{}'.format(nrows))

    pbar = tqdm(total=nrows * ncols)

    for row in range(1, nrows + 1):
        for col in range(1, ncols + 1):
            pbar.update(1)
            if ws.cell(row, col).value != None:
                input = ws.cell(row, col).value
                if type(input) != str:
                    input = str(input)
                try:
                    ws.cell(row, col).value = translator.translate(input, dest='en').text
                except:
                    ws.cell(row, col).value = 'error'
    try:
        wb.save(out_home)
        print('Result save success!')
    except:
        print("Error!")
        sys.exit()


if __name__ == '__main__':
    file_path = '../data/q_list.xlsx'
    out_path = '../data/q_list_en_3.xlsx'
    Google_Translate(file_path, out_path)
